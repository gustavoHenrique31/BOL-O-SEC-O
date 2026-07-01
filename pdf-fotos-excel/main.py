# main.py — Aplicação PDF → Fotos Excel
# Frontend: mantido conforme arquivo TXT fornecido
# Backend: corrigido — leitura Excel + inserção de imagens

import base64
import io
import json
import os
import re
import tempfile
import unicodedata
from typing import Dict, List

from fastapi import FastAPI, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from starlette.background import BackgroundTask

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps

# ────────────────────────────────────────────────────────────────
#  APP
# ────────────────────────────────────────────────────────────────

app = FastAPI(title="PDF → Fotos Excel")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ────────────────────────────────────────────────────────────────
#  UTILITÁRIOS
# ────────────────────────────────────────────────────────────────

def safe_int(v, default: int) -> int:
    try:
        return int(v)
    except Exception:
        return default


def normalize(value) -> str:
    if value is None:
        return ""
    t = unicodedata.normalize("NFKD", str(value).strip())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).lower()


def is_foto_col(value) -> bool:
    """
    Reconhece colunas de foto:
    Foto_1, Foto 1, FOTO_2, foto-3, Foto, etc.
    """
    return bool(re.match(r"^foto[\s_\-]*\d*$", normalize(value)))


def decode_b64(data: str) -> bytes:
    """
    Aceita 'data:image/jpeg;base64,XXX' ou somente 'XXX'.
    """
    if not data:
        raise ValueError("Imagem vazia")
    if "," in data:
        data = data.split(",", 1)[1]
    return base64.b64decode(data)


def resize_image(raw: bytes, max_w: int, max_h: int, quality: int = 88):
    """
    Redimensiona mantendo proporção.
    Retorna (buffer, w_final, h_final).
    """
    with PILImage.open(io.BytesIO(raw)) as img:
        img = ImageOps.exif_transpose(img)
        if img.mode != "RGB":
            img = img.convert("RGB")
        img.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality)
        buf.seek(0)
        return buf, img.width, img.height


def get_temp_filename(suffix: str = ".xlsx") -> str:
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    return path


# ────────────────────────────────────────────────────────────────
#  FRONTEND (HTML + JS)
# ────────────────────────────────────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="pt-BR"

> <head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>PDF → Fotos Excel</title>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.min.js"></script>
  <style>
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    body{font-family:'Segoe UI',sans-serif;background:#0f1117;color:#e2e8f0;min-height:100vh}
    header{background:linear-gradient(135deg,#1e40af,#7c3aed);padding:20px 32px;display:flex;align-items:center;gap:14px;box-shadow:0 4px 20px rgba(0,0,0,.4)}
    header .logo{font-size:2rem}
    header h1{font-size:1.35rem;font-weight:700}
    header p{font-size:.8rem;opacity:.8;margin-top:2px}
    .container{max-width:1300px;margin:0 auto;padding:28px 20px}
    .backend-status{display:flex;align-items:center;gap:10px;background:#1e2130;border-radius:10px;padding:10px 16px;margin-bottom:20px;font-size:.82rem}
    .dot{width:10px;height:10px;border-radius:50%;background:#ef4444;flex-shrink:0;transition:background .3s}
    .dot.online{background:#10b981}
    .backend-status span{color:#94a3b8}
    .backend-status strong{color:#e2e8f0}
    .upload-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
    @media(max-width:700px){.upload-grid{grid-template-columns:1fr}}
    .upload-zone{background:#1e2130;border:2px dashed #3b4262;border-radius:14px;padding:28px 20px;text-align:center;transition:all .25s}
    .upload-zone.drag-over,.upload-zone:hover{border-color:#6366f1;background:#1a1d3a}
    .upload-zone label{display:block;cursor:pointer}
    .upload-zone input[type=file]{display:none}
    .uz-icon{font-size:2.4rem;margin-bottom:8px}
    .upload-zone h2{font-size:.95rem;font-weight:700;margin-bottom:5px}
    .upload-zone p{font-size:.76rem;color:#94a3b8}
    .fn{margin-top:10px;font-size:.78rem;color:#a5b4fc;font-weight:600;word-break:break-all;display:none}
    .fn.visible{display:block}
    .config-panel{background:#1e2130;border-radius:14px;padding:18px 20px;margin-bottom:18px;display:grid;grid-template-columns:repeat(3,1fr);gap:14px;align-items:end}
    @media(max-width:900px){.config-panel{grid-template-columns:1fr 1fr}}
    .field label{display:block;font-size:.72rem;color:#94a3b8;margin-bottom:4px;font-weight:600;text-transform:uppercase;letter-spacing:.4px}
    .field input,.field select{width:100%;background:#0f1117;border:1.5px solid #3b4262;border-radius:8px;padding:8px 11px;color:#e2e8f0;font-size:.86rem;outline:none;transition:border-color .2s}
    .field input:focus,.field select:focus{border-color:#6366f1}
    .adv-panel{background:#1a1d2e;border:1px solid #2d3155;border-radius:12px;padding:16px 18px;margin-bottom:18px}
    .adv-panel h4{font-size:.76rem;font-weight:700;color:#a5b4fc;text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px;display:flex;align-items:center;gap:8px}
    .adv-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}
    @media(max-width:900px){.adv-grid{grid-template-columns:1fr 1fr}}
    .toggle-row{display:flex;align-items:center;gap:9px;background:#0f1117;border-radius:8px;padding:8px 11px;border:1px solid #2d3155}
    .toggle-row label{font-size:.74rem;color:#94a3b8;cursor:pointer;flex:1}
    .toggle-row input[type=checkbox]{width:14px;height:14px;accent-color:#6366f1;cursor:pointer}
    .info-banner{display:none;border-radius:10px;padding:11px 15px;margin-bottom:16px;font-size:.8rem;gap:9px;align-items:flex-start}
    .info-banner.visible{display:flex}
    .info-banner.ok{background:#1a2744;border:1px solid #3b5998;color:#93c5fd}
    .info-banner.ok strong{color:#bfdbfe}
    .info-banner.warn{background:#2d1f00;border:1px solid #92400e;color:#fcd34d}
    .info-banner.err{background:#2d0000;border:1px solid #7f1d1d;color:#fca5a5}
    .btn-extract{width:100%;padding:14px;border-radius:11px;background:linear-gradient(135deg,#6366f1,#8b5cf6);color:#fff;font-size:.96rem;font-weight:700;border:none;cursor:pointer;transition:all .25s;display:flex;align-items:center;justify-content:center;gap:9px;margin-bottom:20px}
    .btn-extract:hover{opacity:.88;transform:translateY(-2px)}
    .btn-extract:disabled{opacity:.35;cursor:not-allowed;transform:none}
    .progress-wrap{display:none;background:#1e2130;border-radius:12px;padding:14px 18px;margin-bottom:16px}
    .progress-wrap.visible{display:block}
    .prog-top{display:flex;justify-content:space-between;font-size:.78rem;color:#94a3b8;margin-bottom:6px}
    .prog-bg{height:9px;background:#0f1117;border-radius:99px;overflow:hidden;margin-bottom:4px}
    .prog-fill{height:100%;width:0%;border-radius:99px;background:linear-gradient(90deg,#6366f1,#8b5cf6);transition:width .3s}
    .sub-bg{height:4px;background:#0f1117;border-radius:99px;overflow:hidden}
    .sub-fill{height:100%;width:0%;border-radius:99px;background:linear-gradient(90deg,#10b981,#059669);transition:width .2s}
    .log-box{background:#0a0c12;border:1px solid #1e2130;border-radius:9px;padding:8px 12px;font-size:.7rem;color:#64748b;max-height:110px;overflow-y:auto;margin-bottom:16px;display:none;font-family:monospace;line-height:1.7}
    .log-box.visible{display:block}
    .lok{color:#10b981}.lwarn{color:#f59e0b}.lerr{color:#ef4444}.linfo{color:#818cf8}
    .stats-bar{display:none;gap:9px;flex-wrap:wrap;margin-bottom:16px}
    .stats-bar.visible{display:flex}
    .stat-chip{background:#1e2130;border-radius:8px;padding:6px 14px;font-size:.77rem;display:flex;align-items:center;gap:7px}
    .stat-chip .sv{font-weight:700;color:#a5b4fc}
    .assign-panel{background:#1e2130;border-radius:14px;padding:18px 20px;margin-bottom:20px;display:none}
    .assign-panel.visible{display:block}
    .assign-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;flex-wrap:wrap;gap:10px}
    .assign-header h3{font-size:.95rem;font-weight:700}
    .assign-actions{display:flex;gap:8px;flex-wrap:wrap}
    .col-tabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px;padding-bottom:12px;border-bottom:1px solid #2d3155}
    .col-tab{padding:5px 14px;border-radius:99px;border:1.5px solid #3b4262;background:#0f1117;color:#94a3b8;font-size:.76rem;font-weight:600;cursor:pointer;transition:all .2s}
    .col-tab.active{border-color:#6366f1;color:#fff;background:#4f46e5}
    .col-tab.has-photos{border-color:#10b981;color:#10b981}
    .col-tab.active.has-photos{background:#059669;border-color:#059669;color:#fff}
    .assign-instr{font-size:.75rem;color:#64748b;margin-bottom:10px}
    .assign-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(130px,1fr));gap:9px}
    .assign-card{border-radius:10px;overflow:hidden;border:2px solid #2d3155;cursor:pointer;transition:all .2s;background:#0f1117;position:relative}
    .assign-card:hover{transform:translateY(-2px);box-shadow:0 6px 18px rgba(0,0,0,.4)}
    .assign-card.assigned{border-color:#6366f1;background:#1a1d3a}
    .assign-card .ac-thumb{aspect-ratio:1;display:flex;align-items:center;justify-content:center;overflow:hidden;background:#0a0c12}
    .assign-card .ac-thumb img{max-width:100%;max-height:100%;object-fit:contain}
    .assign-card .ac-col-tag{position:absolute;top:5px;right:5px;font-size:.6rem;font-weight:700;padding:2px 7px;border-radius:99px;background:#6366f1;color:#fff;display:none}
    .assign-card.assigned .ac-col-tag{display:block}
    .assign-card .ac-foot{padding:4px 7px;display:flex;justify-content:space-between}
    .assign-card .ac-idx{font-size:.64rem;color:#64748b;font-weight:600}
    .assign-card .ac-pg{font-size:.6rem;color:#475569}
    .assign-summary{margin-top:14px;padding-top:12px;border-top:1px solid #2d3155;display:flex;flex-wrap:wrap;gap:8px}
    .sum-chip{background:#0f1117;border-radius:8px;padding:5px 12px;font-size:.75rem;color:#94a3b8;border:1px solid #2d3155}
    .sum-chip strong{color:#a5b4fc}
    .gallery-header{font-size:.95rem;font-weight:700;margin-bottom:11px;display:flex;align-items:center;gap:9px}
    .badge{background:#6366f1;color:#fff;font-size:.7rem;font-weight:700;padding:2px 9px;border-radius:99px}
    .page-filters{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:11px}
    .pf-btn{padding:3px 11px;border-radius:99px;border:1.5px solid #3b4262;background:#1e2130;color:#94a3b8;font-size:.72rem;font-weight:600;cursor:pointer;transition:all .2s}
    .pf-btn.active,.pf-btn:hover{border-color:#6366f1;color:#a5b4fc;background:#1a1d3a}
    .sel-bar{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:14px}
    .sel-bar .spacer{flex:1}
    .btn-sm{display:inline-flex;align-items:center;gap:5px;padding:7px 13px;border-radius:8px;font-size:.78rem;font-weight:600;cursor:pointer;border:none;transition:all .2s;white-space:nowrap}
    .btn-outline{background:transparent;border:1.5px solid #3b4262;color:#e2e8f0}
    .btn-outline:hover{border-color:#6366f1;color:#a5b4fc}
    .btn-green{background:linear-gradient(135deg,#059669,#10b981);color:#fff}
    .btn-green:hover{opacity:.88;transform:translateY(-1px)}
    .btn-green:disabled{opacity:.35;cursor:not-allowed;transform:none}
    .sel-info{font-size:.78rem;color:#94a3b8;background:#1e2130;padding:5px 11px;border-radius:8px}
    .gallery{display:grid;grid-template-columns:repeat(auto-fill,minmax(155px,1fr));gap:10px;margin-bottom:28px}
    .img-card{background:#1e2130;border-radius:10px;overflow:hidden;border:2px solid transparent;cursor:pointer;transition:all .2s;position:relative}
    .img-card:hover{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,.4)}
    .img-card.selected{border-color:#6366f1;box-shadow:0 0 0 3px rgba(99,102,241,.25)}
    .img-card.rejected{opacity:.25;filter:grayscale(80%)}
    .thumb{position:relative;aspect-ratio:1;background:#0a0c12;display:flex;align-items:center;justify-content:center;overflow:hidden}
    .thumb img{max-width:100%;max-height:100%;object-fit:contain;display:block}
    .chk{position:absolute;top:6px;right:6px;width:21px;height:21px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.76rem;font-weight:700;z-index:2}
    .img-card.selected .chk{background:#6366f1;color:#fff}
    .img-card:not(.selected):not(.rejected) .chk{background:rgba(255,255,255,.1);color:#94a3b8}
    .img-card.rejected .chk{background:#dc2626;color:#fff}
    .pg-badge{position:absolute;bottom:4px;left:4px;background:rgba(0,0,0,.75);color:#94a3b8;font-size:.6rem;padding:1px 6px;border-radius:99px;z-index:2}
    .sq-badge{position:absolute;top:6px;left:6px;background:rgba(0,0,0,.75);color:#e2e8f0;font-size:.6rem;padding:1px 6px;border-radius:99px;font-weight:700;z-index:2}
    .card-foot{padding:5px 8px;display:flex;justify-content:space-between}
    .card-foot .ci{font-size:.67rem;color:#64748b;font-weight:600}
    .card-foot .cs{font-size:.62rem;color:#475569}
    .card-acts{display:flex;gap:3px;padding:0 8px 7px}
    .mb{flex:1;padding:4px;border-radius:6px;border:none;font-size:.63rem;font-weight:600;cursor:pointer;transition:all .15s}
    .mb.ok{background:#1a3a2a;color:#10b981}.mb.ok:hover{background:#059669;color:#fff}
    .mb.no{background:#3a1a1a;color:#f87171}.mb.no:hover{background:#dc2626;color:#fff}
    .empty{text-align:center;padding:48px 20px;color:#475569}
    .empty .ei{font-size:3.2rem;margin-bottom:12px}
    .empty p{font-size:.86rem}
    .modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.82);z-index:1000;align-items:center;justify-content:center}
    .modal-bg.visible{display:flex}
    .modal{background:#1e2130;border-radius:18px;padding:34px 40px;max-width:440px;width:92%;text-align:center;box-shadow:0 20px 60px rgba(0,0,0,.6)}
    .modal .mi{font-size:2.6rem;margin-bottom:12px}
    .modal h3{font-size:1.05rem;font-weight:700;margin-bottom:5px}
    .modal .ms{font-size:.78rem;color:#94a3b8;margin-bottom:16px}
    .mpb{height:11px;background:#0f1117;border-radius:99px;overflow:hidden;margin-bottom:7px}
    .mpf{height:100%;width:0%;border-radius:99px;background:linear-gradient(90deg,#6366f1,#10b981);transition:width .4s}
    .mlb{font-size:.78rem;color:#94a3b8;margin-bottom:3px}
    .mcnt{font-size:.98rem;font-weight:700;color:#a5b4fc}
    .ok-banner{display:none;background:linear-gradient(135deg,#064e3b,#065f46);border:1px solid #10b981;border-radius:11px;padding:13px 17px;margin-bottom:16px;align-items:center;gap:12px}
    .ok-banner.visible{display:flex}
    .ok-banner .obi{font-size:1.6rem}
    .ok-banner h4{font-size:.9rem;font-weight:700;color:#6ee7b7}
    .ok-banner p{font-size:.76rem;color:#a7f3d0;margin-top:2px}
    #toast{position:fixed;bottom:22px;right:22px;background:#1e2130;color:#e2e8f0;padding:11px 17px;border-radius:10px;font-size:.82rem;font-weight:500;box-shadow:0 8px 28px rgba(0,0,0,.5);border-left:4px solid #6366f1;transform:translateY(80px);opacity:0;transition:all .3s;z-index:9999;max-width:330px}
    #toast.show{transform:translateY(0);opacity:1}
    #toast.success{border-left-color:#10b981}
    #toast.error{border-left-color:#ef4444}
    #toast.warn{border-left-color:#f59e0b}
  </style>
</head>
<body>
<header>
  <div class="logo">🖼️</div>
  <div>
    <h1>PDF → Fotos Excel</h1>
    <p>Extraia imagens do PDF e insira nas colunas Foto_X do template</p>
  </div>
</header>
<div class="container"

>   <div class="backend-status" id="backendStatus"

>     <div class="dot" id="backendDot"></div>
    <div>
      <strong id="backendStatusText">Verificando servidor...</strong>
      <span id="backendStatusSub"

> — aguarde</span>
    </div>
  </div>
  <div class="upload-grid"

>     <div class="upload-zone" id="pdfDZ"

>       <label for="pdfInput"

>         <div class="uz-icon">📄</div>
        <h2>Clique ou arraste o PDF</h2>
        <p>Imagens extraídas automaticamente</p>
        <div class="fn" id="pdfFN"></div>
      </label>
      <input type="file" id="pdfInput" accept=".pdf"/>
    </div>
    <div class="upload-zone" id="excelDZ"

>       <label for="excelInput"

>         <div class="uz-icon">📊</div>
        <h2>Clique ou arraste o template Excel</h2>
        <p>Detecta colunas Foto_X automaticamente</p>
        <div class="fn" id="excelFN"></div>
      </label>
      <input type="file" id="excelInput" accept=".xlsx"/>
    </div>
  </div>
  <div class="info-banner" id="infoBanner"><span>🎯</span><div id="infoBannerText"></div></div>
  <div class="config-panel"

>     <div class="field"

>       <label>📋 Aba de destino</label>
      <select id="sheetSelect" onchange="onSheetChange()"><option value="">Carregue o Excel</option></select>
    </div>
    <div class="field"

>       <label>🔢 Linha do cabeçalho</label>
      <input type="number" id="headerRowInput" value="1" min="1" max="50" onchange="onSheetChange()"/>
    </div>
    <div class="field"

>       <label>📐 Escala renderização PDF</label>
      <select id="scaleSelect"

>         <option value="1">1x</option>
        <option value="1.5" selected>1.5x ✅</option>
        <option value="2">2x</option>
        <option value="3">3x</option>
      </select>
    </div>
  </div>
  <div class="adv-panel"

>     <h4>⚙️ Opções</h4>
    <div class="adv-grid"

>       <div class="field"><label>Tam. mínimo (px)</label><input type="number" id="minSize" value="40" min="5" max="300"/></div>
      <div class="field"><label>Dedup (%)</label><input type="number" id="overlapThresh" value="85" min="0" max="100"/></div>
      <div class="field"><label>Qualidade JPEG</label><input type="number" id="jpegQ" value="92" min="50" max="100"/></div>
      <div class="field"><label>Margem recorte (px)</label><input type="number" id="cropPad" value="3" min="0" max="20"/></div>
      <div class="toggle-row"><input type="checkbox" id="chkClean" checked/><label for="chkClean">🧹 Remover logos</label></div>
      <div class="toggle-row"><input type="checkbox" id="chkDedup" checked/><label for="chkDedup">🔁 Dedup IoU</label></div>
      <div class="toggle-row"><input type="checkbox" id="chkFallback" checked/><label for="chkFallback">🛡️ Fallback página</label></div>
      <div class="toggle-row"><input type="checkbox" id="chkAutoAssign"/><label for="chkAutoAssign">⚡ Auto-distribuir</label></div>
    </div>
  </div>
  <button class="btn-extract" id="btnExtract" disabled onclick="extractImages()">🔍 Extrair Imagens do PDF</button>
  <div class="progress-wrap" id="progressWrap"

>     <div class="prog-top"><span id="progText">Aguardando...</span><span id="progPct">0%</span></div>
    <div class="prog-bg"><div class="prog-fill" id="progFill"></div></div>
    <div class="sub-bg"><div class="sub-fill" id="subFill"></div></div>
  </div>
  <div class="log-box" id="logBox"></div>
  <div class="ok-banner" id="okBanner"

>     <div class="obi">✅</div>
    <div><h4 id="okTitle">Excel gerado!</h4><p id="okDesc"></p></div>
  </div>
  <div class="stats-bar" id="statsBar"

>     <div class="stat-chip">📄 Páginas: <span class="sv" id="stPages">0</span></div>
    <div class="stat-chip">🖼️ Extraídas: <span class="sv" id="stImgs">0</span></div>
    <div class="stat-chip">✅ Sel: <span class="sv" id="stSel">0</span></div>
    <div class="stat-chip">❌ Excl: <span class="sv" id="stRej">0</span></div>
  </div>
  <div class="assign-panel" id="assignPanel"

>     <div class="assign-header"

>       <h3>🗂️ Atribuição por Coluna Foto</h3>
      <div class="assign-actions"

>         <button class="btn-sm btn-outline" onclick="clearAllAssignments()">🗑️ Limpar</button>
        <button class="btn-sm btn-outline" onclick="autoAssignAll()">⚡ Auto</button>
        <button class="btn-sm btn-green" id="btnExport" disabled onclick="exportExcel()">📥 Exportar Excel</button>
      </div>
    </div>
    <div class="col-tabs" id="colTabs"></div>
    <div class="assign-instr" id="assignInstr">Selecione uma coluna e clique nas fotos</div>
    <div class="assign-grid" id="assignGrid"></div>
    <div class="assign-summary" id="assignSummary"></div>
  </div>
  <div id="gallerySection" style="display:none"

>     <div class="gallery-header">🖼️ Imagens Extraídas <span class="badge" id="totalBadge">0</span></div>
    <div class="page-filters" id="pageFilters"></div>
    <div class="sel-bar"

>       <button class="btn-sm btn-outline" onclick="selAll(true)">✅ Todas</button>
      <button class="btn-sm btn-outline" onclick="selAll(false)">⬜ Nenhuma</button>
      <button class="btn-sm btn-outline" onclick="selPage(true)">📄✅</button>
      <button class="btn-sm btn-outline" onclick="selPage(false)">📄⬜</button>
      <div class="sel-info"><span id="selCount">0</span>/<span id="totalCount">0</span></div>
      <div class="spacer"></div>
    </div>
    <div class="gallery" id="gallery"></div>
  </div>
  <div class="empty" id="emptyState"

>     <div class="ei">📂</div>
    <p>Carregue um PDF e um template Excel para começar.</p>
  </div>
</div>
<div class="modal-bg" id="modalBg"

>   <div class="modal"

>     <div class="mi" id="mIcon">⚙️</div>
    <h3 id="mTitle">Gerando Excel</h3>
    <div class="ms" id="mSub">Aguarde...</div>
    <div class="mpb"><div class="mpf" id="mPF"></div></div>
    <div class="mlb" id="mLbl">Iniciando...</div>
    <div class="mcnt" id="mCnt">0/0</div>
  </div>
</div>
<div id="toast"></div>
<script>
pdfjsLib.GlobalWorkerOptions.workerSrc='https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.worker.min.js';
const $=id=>document.getElementById(id);
const BACKEND='/api';
let pdfFile=null,excelFile=null;
let extractedImages=[],photoColumns=[],assignments={};
let activeFilter='all',activeAssignCol=null,backendOnline=false;

function toast(msg,type='info'){const t=$('toast');t.textContent=msg;t.className=`show ${type}`;clearTimeout(t._t);t._t=setTimeout(()=>t.className='',4200)}
function log(msg,type='info'){const b=$('logBox');b.classList.add('visible');const d=document.createElement('div');d.className=`l${type}`;d.textContent=`[${new Date().toLocaleTimeString()}] ${msg}`;b.appendChild(d);b.scrollTop=b.scrollHeight}
function setP(pct,lbl){$('progFill').style.width=pct+'%';$('progPct').textContent=pct+'%';$('progText').textContent=lbl}
function setSub(pct){$('subFill').style.width=pct+'%'}
function setM(pct,lbl,cnt){$('mPF').style.width=pct+'%';$('mLbl').textContent=lbl;$('mCnt').textContent=cnt}
function yf(){return new Promise(r=>requestAnimationFrame(()=>setTimeout(r,0)))}
function showBanner(msg,type='ok'){const b=$('infoBanner');b.className=`info-banner ${type} visible`;$('infoBannerText').innerHTML=msg}
function hideBanner(){$('infoBanner').classList.remove('visible')}

async function checkBackend(){
  try{
    const r=await fetch(BACKEND+'/',{signal:AbortSignal.timeout(5000)});
    const d=await r.json();
    if(r.ok&&d.status==='ok'){
      backendOnline=true;$('backendDot').classList.add('online');
      $('backendStatusText').textContent='Servidor online';
      $('backendStatusSub').textContent=' — pronto para exportar';
    }else throw new Error();
  }catch{
    backendOnline=false;$('backendDot').classList.remove('online');
    $('backendStatusText').textContent='Servidor offline';
    $('backendStatusSub').textContent=' — aguardando conexão';
  }
}
checkBackend();setInterval(checkBackend,15000);

$('pdfInput').addEventListener('change',function(){
  if(!this.files[0])return;
  if(!this.files[0].name.toLowerCase().endsWith('.pdf')){toast('⚠️ .pdf apenas','warn');return}
  pdfFile=this.files[0];$('pdfFN').textContent='📄 '+pdfFile.name;$('pdfFN').classList.add('visible');
  checkReady();toast(`📄 ${pdfFile.name}`,'success');
});
$('excelInput').addEventListener('change',function(){if(this.files[0])loadExcelInfo(this.files[0])});

['pdfDZ','excelDZ'].forEach(id=>{
  const z=$(id);
  z.addEventListener('dragover',e=>{e.preventDefault();e.stopPropagation();z.classList.add('drag-over')});
  z.addEventListener('dragleave',e=>{e.stopPropagation();z.classList.remove('drag-over')});
  z.addEventListener('drop',e=>{
    e.preventDefault();e.stopPropagation();z.classList.remove('drag-over');
    const f=e.dataTransfer.files[0];if(!f)return;
    if(id==='pdfDZ'){
      if(!f.name.toLowerCase().endsWith('.pdf')){toast('⚠️ .pdf apenas','warn');return}
      pdfFile=f;$('pdfFN').textContent='📄 '+f.name;$('pdfFN').classList.add('visible');
      checkReady();toast(`📄 ${f.name}`,'success');
    }else{
      if(!f.name.toLowerCase().endsWith('.xlsx')){toast('⚠️ .xlsx apenas','warn');return}
      loadExcelInfo(f);
    }
  });
});

async function loadExcelInfo(f){
  excelFile=f;$('excelFN').textContent='📊 '+f.name;$('excelFN').classList.add('visible');
  checkReady();toast(`📊 ${f.name}`,'success');
  if(!backendOnline){showBanner('⚠️ Servidor offline.','warn');return}
  try{
    const fd=new FormData();fd.append('template',f);
    const r=await fetch(BACKEND+'/info-abas',{method:'POST',body:fd});
    const d=await r.json();
    const sel=$('sheetSelect');sel.innerHTML='';
    (d.sheets||[]).forEach(n=>{const o=document.createElement('option');o.value=n;o.textContent=n;sel.appendChild(o)});
    log(`Excel: ${(d.sheets||[]).length} aba(s)`,'ok');
    await loadPhotoColumns();
  }catch(err){showBanner(`❌ Erro: ${err.message}`,'err')}
}

async function onSheetChange(){await loadPhotoColumns()}

function idxToCol(idx){const A='ABCDEFGHIJKLMNOPQRSTUVWXYZ';let s='';let n=idx;while(n>0){n--;s=A[n%26]+s;n=Math.floor(n/26)}return s}

async function loadPhotoColumns(){
  const sn=$('sheetSelect').value,hr=parseInt($('headerRowInput').value)||1;
  hideBanner();photoColumns=[];assignments={};
  if(!excelFile||!sn)return;
  if(!backendOnline){showBanner('⚠️ Servidor offline.','warn');return}
  try{
    const fd=new FormData();fd.append('template',excelFile);fd.append('sheet_name',sn);fd.append('header_row',String(hr));
    const r=await fetch(BACKEND+'/info-colunas',{method:'POST',body:fd});
    const d=await r.json();
    if(d.error){showBanner(`❌ ${d.error}`,'err');return}
    photoColumns=(d.columns||[]).map(c=>({name:c.name,colNum:c.col,letter:c.letter||idxToCol(c.col)}));
    photoColumns.forEach(c=>{assignments[c.colNum]=[]});
    if(!photoColumns.length){showBanner('⚠️ Nenhuma coluna Foto_X.','warn');return}
    const preview=photoColumns.slice(0,6).map(c=>`<strong>${c.name}</strong>(${c.letter})`).join(' ');
    const more=photoColumns.length>6?` +${photoColumns.length-6}`:'';
    showBanner(`✅ <strong>${photoColumns.length}</strong> coluna(s): ${preview}${more}`,'ok');
    log(`Colunas: ${photoColumns.map(c=>c.name).join(', ')}`,'ok');
    if(extractedImages.length>0)buildAssignPanel();
  }catch(err){showBanner(`❌ ${err.message}`,'err')}
}

function checkReady(){$('btnExtract').disabled=!(pdfFile&&excelFile)}

async function extractImages(){
  if(!pdfFile)return;
  extractedImages=[];assignments={};photoColumns.forEach(c=>{assignments[c.colNum]=[]});
  $('gallery').innerHTML='';$('logBox').innerHTML='';$('logBox').classList.remove('visible');
  $('gallerySection').style.display='none';$('assignPanel').classList.remove('visible');
  $('emptyState').style.display='none';$('statsBar').classList.remove('visible');
  $('okBanner').classList.remove('visible');$('pageFilters').innerHTML='';activeFilter='all';
  const pw=$('progressWrap');pw.classList.add('visible');setP(0,'Carregando PDF...');setSub(0);
  const scale=parseFloat($('scaleSelect').value)||1.5;
  const minSize=parseInt($('minSize').value)||40;
  const overlapT=parseInt($('overlapThresh').value)||85;
  const jpegQ=(parseInt($('jpegQ').value)||92)/100;
  const cropPad=parseInt($('cropPad').value)||3;
  const doDedup=$('chkDedup').checked;
  const doFallback=$('chkFallback').checked;
  try{
    const buf=await pdfFile.arrayBuffer();
    const pdf=await pdfjsLib.getDocument({data:buf}).promise;
    const total=pdf.numPages;log(`PDF: ${total} páginas`,'ok');
    $('stPages').textContent=total;
    for(let p=1;p<=total;p++){
      setP(Math.round(((p-1)/total)*95),`Página ${p}/${total}...`);setSub(0);await yf();
      const page=await pdf.getPage(p);
      const vp=page.getViewport({scale});
      const canvas=document.createElement('canvas');
      canvas.width=Math.round(vp.width);canvas.height=Math.round(vp.height);
      await page.render({canvasContext:canvas.getContext('2d'),viewport:vp}).promise;
      setSub(40);
      const ops=await page.getOperatorList();
      const regions=detectRegions(ops,page,vp);
      const unique=doDedup?dedup(regions,overlapT/100):regions;
      setSub(70);let cnt=0;
      for(const reg of unique){
        const x0=Math.max(0,Math.round(reg.x)-cropPad),y0=Math.max(0,Math.round(reg.y)-cropPad);
        const x1=Math.min(canvas.width,Math.round(reg.x+reg.w)+cropPad);
        const y1=Math.min(canvas.height,Math.round(reg.y+reg.h)+cropPad);
        const cw=x1-x0,ch=y1-y0;if(cw<minSize||ch<minSize)continue;
        const crop=document.createElement('canvas');crop.width=cw;crop.height=ch;
        crop.getContext('2d').drawImage(canvas,x0,y0,cw,ch,0,0,cw,ch);
        cnt++;extractedImages.push({dataUrl:crop.toDataURL('image/jpeg',jpegQ),w:cw,h:ch,page:p,seq:cnt,idx:extractedImages.length,selected:true});
      }
      if(cnt===0&&doFallback){
        log(`P${p}: fallback`,'warn');
        extractedImages.push({dataUrl:canvas.toDataURL('image/jpeg',jpegQ),w:canvas.width,h:canvas.height,page:p,seq:1,idx:extractedImages.length,selected:true});cnt=1;
      }
      setSub(100);log(`P${p}: ${cnt} foto(s)`,'ok');
    }
    setP(100,'Concluído!');setTimeout(()=>pw.classList.remove('visible'),900);
    $('stImgs').textContent=extractedImages.length;
    if(!extractedImages.length){toast('⚠️ Nenhuma imagem.','warn');$('emptyState').style.display='block';return}
    extractedImages=cleanImages(extractedImages);extractedImages.forEach((img,i)=>img.idx=i);
    if(!extractedImages.length){toast('⚠️ Todas removidas.','warn');$('emptyState').style.display='block';return}
    buildGallery();buildAssignPanel();
    if($('chkAutoAssign').checked&&photoColumns.length>0)autoAssignAll();
    toast(`✅ ${extractedImages.length} imagens!`,'success');
  }catch(err){pw.classList.remove('visible');toast('❌ '+err.message,'error');log('ERRO: '+err.message,'err');console.error(err)}
}

function detectRegions(ops,page,vp){
  const regions=[],stack=[];let ctm=[1,0,0,1,0,0];const OPS=pdfjsLib.OPS;
  for(let i=0;i<ops.fnArray.length;i++){
    const fn=ops.fnArray[i],args=ops.argsArray[i];
    switch(fn){
      case OPS.save:stack.push([...ctm]);break;
      case OPS.restore:if(stack.length)ctm=stack.pop();break;
      case OPS.transform:ctm=mm(ctm,args);break;
      case OPS.paintFormXObjectBegin:stack.push([...ctm]);if(args&&args[1])ctm=mm(ctm,args[1]);break;
      case OPS.paintFormXObjectEnd:if(stack.length)ctm=stack.pop();break;
      case OPS.paintImageXObject:case OPS.paintJpegXObject:case OPS.paintImageMaskXObject:regions.push(bbox(ctm,page,vp));break;
    }
  }
  return regions;
}
function bbox(ctm,page,vp){
  const pts=[[0,0],[1,0],[1,1],[0,1]].map(([px,py])=>{const ux=ctm[0]*px+ctm[2]*py+ctm[4],uy=ctm[1]*px+ctm[3]*py+ctm[5];return vp.convertToViewportPoint(ux,uy)});
  const xs=pts.map(p=>p[0]),ys=pts.map(p=>p[1]);
  return{x:Math.min(...xs),y:Math.min(...ys),w:Math.max(...xs)-Math.min(...xs),h:Math.max(...ys)-Math.min(...ys)};
}
function mm(a,b){return[a[0]*b[0]+a[2]*b[1],a[1]*b[0]+a[3]*b[1],a[0]*b[2]+a[2]*b[3],a[1]*b[2]+a[3]*b[3],a[0]*b[4]+a[2]*b[5]+a[4],a[1]*b[4]+a[3]*b[5]+a[5]]}
function dedup(r,t){const k=[];for(const x of r){if(!k.some(y=>iou(y,x)>=t))k.push(x)}return k}
function iou(a,b){const x0=Math.max(a.x,b.x),y0=Math.max(a.y,b.y),x1=Math.min(a.x+a.w,b.x+b.w),y1=Math.min(a.y+a.h,b.y+b.h);const i=Math.max(0,x1-x0)*Math.max(0,y1-y0);if(!i)return 0;return i/(a.w*a.h+b.w*b.h-i)}
function cleanImages(images){
  if(!$('chkClean').checked)return images;
  // aqui você pode manter sua lógica antiga de remoção de logos/selo
  return images;
}

function buildGallery(){
  $('gallerySection').style.display='block';$('emptyState').style.display='none';
  const g=$('gallery');g.innerHTML='';
  const pages=[...new Set(extractedImages.map(i=>i.page))].sort((a,b)=>a-b);
  const pf=$('pageFilters');pf.innerHTML='';
  const btnAll=document.createElement('button');btnAll.className='pf-btn active';btnAll.textContent='Todas';btnAll.onclick=()=>{activeFilter='all';updateGallery();};
  pf.appendChild(btnAll);
  pages.forEach(p=>{
    const b=document.createElement('button');b.className='pf-btn';b.textContent='P'+p;b.onclick=()=>{activeFilter=p;updateGallery();};pf.appendChild(b);
  });
  $('totalBadge').textContent=extractedImages.length;
  $('totalCount').textContent=extractedImages.length;
  $('statsBar').classList.add('visible');
  updateGallery();
}

function updateGallery(){
  const g=$('gallery');g.innerHTML='';
  let sel=0,rej=0;
  extractedImages.forEach(img=>{
    if(activeFilter!=='all'&&img.page!==activeFilter)return;
    const card=document.createElement('div');card.className='img-card';if(img.selected)card.classList.add('selected');
    card.dataset.idx=img.idx;
    const thumb=document.createElement('div');thumb.className='thumb';
    const im=document.createElement('img');im.src=img.dataUrl;thumb.appendChild(im);
    const chk=document.createElement('div');chk.className='chk';chk.textContent=img.selected?'✓':'';
    const pg=document.createElement('div');pg.className='pg-badge';pg.textContent='P'+img.page;
    thumb.appendChild(chk);thumb.appendChild(pg);
    card.appendChild(thumb);
    const foot=document.createElement('div');foot.className='card-foot';
    const ci=document.createElement('div');ci.className='ci';ci.textContent='#'+img.idx;
    const cs=document.createElement('div');cs.className='cs';cs.textContent=img.w+'x'+img.h;
    foot.appendChild(ci);foot.appendChild(cs);
    card.appendChild(foot);
    card.onclick=()=>{img.selected=!img.selected;updateGallery();updateStats();};
    g.appendChild(card);
    if(img.selected)sel++;else rej++;
  });
  $('selCount').textContent=sel;
  $('stSel').textContent=sel;
  $('stRej').textContent=rej;
}

function updateStats(){
  const sel=extractedImages.filter(i=>i.selected).length;
  const rej=extractedImages.length-sel;
  $('selCount').textContent=sel;
  $('stSel').textContent=sel;
  $('stRej').textContent=rej;
}

function selAll(v){extractedImages.forEach(i=>i.selected=v);updateGallery();updateStats();}
function selPage(v){extractedImages.forEach(i=>{if(i.page===activeFilter||activeFilter==='all')i.selected=v});updateGallery();updateStats();}

function buildAssignPanel(){
  if(!photoColumns.length){$('assignPanel').classList.remove('visible');return}
  $('assignPanel').classList.add('visible');
  const tabs=$('colTabs');tabs.innerHTML='';
  photoColumns.forEach(col=>{
    const b=document.createElement('button');b.className='col-tab';b.textContent=col.name+' ('+col.letter+')';
    b.onclick=()=>{activeAssignCol=col.colNum;updateAssignGrid();document.querySelectorAll('.col-tab').forEach(x=>x.classList.remove('active'));b.classList.add('active');};
    tabs.appendChild(b);
  });
  activeAssignCol=photoColumns[0].colNum;
  tabs.firstChild.classList.add('active');
  updateAssignGrid();
}

function updateAssignGrid(){
  const grid=$('assignGrid');grid.innerHTML='';
  const col=activeAssignCol;if(!col)return;
  extractedImages.forEach(img=>{
    if(!img.selected)return;
    const card=document.createElement('div');card.className='assign-card';
    const thumb=document.createElement('div');thumb.className='ac-thumb';
    const im=document.createElement('img');im.src=img.dataUrl;thumb.appendChild(im);
    card.appendChild(thumb);
    const foot=document.createElement('div');foot.className='ac-foot';
    const idx=document.createElement('div');idx.className='ac-idx';idx.textContent='#'+img.idx;
    const pg=document.createElement('div');pg.className='ac-pg';pg.textContent='P'+img.page;
    foot.appendChild(idx);foot.appendChild(pg);
    card.appendChild(foot);
    const tag=document.createElement('div');tag.className='ac-col-tag';tag.textContent=col;
    card.appendChild(tag);
    card.onclick=()=>toggleAssignment(img.idx,col,card);
    grid.appendChild(card);
  });
  updateAssignSummary();
  $('btnExport').disabled=!hasAssignments();
}

function toggleAssignment(idx,col,cardEl){
  const list=assignments[col]||[];
  const exists=list.find(x=>x.idx===idx);
  if(exists){assignments[col]=list.filter(x=>x.idx!==idx);cardEl.classList.remove('assigned');}
  else{list.push({idx});assignments[col]=list;cardEl.classList.add('assigned');}
  updateAssignSummary();
  $('btnExport').disabled=!hasAssignments();
}

function hasAssignments(){
  return Object.values(assignments).some(lst=>lst&&lst.length>0);
}

function updateAssignSummary(){
  const sum=$('assignSummary');sum.innerHTML='';
  Object.keys(assignments).forEach(col=>{
    const lst=assignments[col]||[];
    if(!lst.length)return;
    const chip=document.createElement('div');chip.className='sum-chip';
    chip.innerHTML='Col '+col+': <strong>'+lst.length+'</strong> foto(s)';
    sum.appendChild(chip);
  });
}

function clearAllAssignments(){
  Object.keys(assignments).forEach(col=>assignments[col]=[]);
  updateAssignGrid();updateAssignSummary();
  $('btnExport').disabled=true;
}

function autoAssignAll(){
  if(!photoColumns.length)return;
  let colIdx=0;
  extractedImages.filter(i=>i.selected).forEach(img=>{
    const col=photoColumns[colIdx%photoColumns.length].colNum;
    const lst=assignments[col]||[];
    if(!lst.find(x=>x.idx===img.idx))lst.push({idx:img.idx});
    assignments[col]=lst;
    colIdx++;
  });
  updateAssignGrid();updateAssignSummary();
  $('btnExport').disabled=!hasAssignments();
}

async function exportExcel(){
  if(!backendOnline){toast('Servidor offline','error');return}
  if(!excelFile){toast('Carregue o Excel','warn');return}
  const sn=$('sheetSelect').value||'';const hr=parseInt($('headerRowInput').value)||1;
  if(!sn){toast('Selecione a aba de destino','warn');return}
  const modalBg=$('modalBg');modalBg.classList.add('visible');setM(0,'Preparando...',`0/${extractedImages.length}`);
  try{
    const fd=new FormData();
    fd.append('template',excelFile);
    fd.append('sheet_name',sn);
    fd.append('header_row',String(hr));
    // construir assignments_json: col -> [{row, dataUrl}]
    const assignmentsPayload={};
    Object.keys(assignments).forEach(col=>{
      const lst=assignments[col]||[];
      assignmentsPayload[col]=lst.map((item,idx)=>({
        row:hr+1+idx, // exemplo: primeira imagem logo abaixo do cabeçalho
        dataUrl:extractedImages[item.idx].dataUrl
      }));
    });
    fd.append('assignments_json',JSON.stringify(assignmentsPayload));
    fd.append('max_width','300');
    fd.append('max_height','300');
    fd.append('jpeg_quality','88');

    setM(30,'Enviando ao servidor...',`0/${extractedImages.length}`);
    const r=await fetch(BACKEND+'/exportar',{method:'POST',body:fd});
    if(!r.ok){
      const err=await r.json().catch(()=>({error:'Erro desconhecido'}));
      throw new Error(err.error||'Falha na exportação');
    }
    setM(80,'Gerando arquivo...',`...`);
    const blob=await r.blob();
    const url=URL.createObjectURL(blob);
    const a=document.createElement('a');
    a.href=url;a.download='fotos_template.xlsx';document.body.appendChild(a);a.click();a.remove();
    URL.revokeObjectURL(url);
    setM(100,'Concluído!',`OK`);
    $('okBanner').classList.add('visible');
    $('okTitle').textContent='Excel gerado!';
    $('okDesc').textContent='Abra o arquivo baixado para conferir as imagens nas colunas Foto_X.';
    toast('✅ Excel gerado com imagens','success');
  }catch(err){
    toast('❌ '+err.message,'error');log('ERRO exportar: '+err.message,'err');
  }finally{
    setTimeout(()=>$('modalBg').classList.remove('visible'),800);
  }
}
</script>
</body>
</html>
"""

# ────────────────────────────────────────────────────────────────
#  ROTAS DE FRONTEND
# ────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def frontend():
  return HTML


# ────────────────────────────────────────────────────────────────
#  ENDPOINT DE SAÚDE DO BACKEND
# ────────────────────────────────────────────────────────────────

@app.get("/api/")
async def healthcheck():
    """
    Frontend usa para checar se o servidor está online.
    """
    return {"status": "ok"}


# ────────────────────────────────────────────────────────────────
#  INFO SOBRE ABAS DO EXCEL
# ────────────────────────────────────────────────────────────────

@app.post("/api/info-abas")
async def info_abas(template: UploadFile = File(...)):
    """
    Recebe o template Excel e retorna a lista de nomes de abas.
    """
    try:
        tmp_path = get_temp_filename(".xlsx")
        with open(tmp_path, "wb") as f:
            f.write(await template.read())

        wb = load_workbook(tmp_path, data_only=True)
        sheets = wb.sheetnames

        def _cleanup():
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        return JSONResponse(
            {"sheets": sheets},
            background=BackgroundTask(_cleanup),
        )
    except Exception as e:
        return JSONResponse(
            {"error": f"Erro ao ler Excel: {e}"},
            status_code=400,
        )


# ────────────────────────────────────────────────────────────────
#  INFO SOBRE COLUNAS FOTO_X
# ────────────────────────────────────────────────────────────────

@app.post("/api/info-colunas")
async def info_colunas(
    template: UploadFile = File(...),
    sheet_name: str = Form(...),
    header_row: int = Form(...),
):
    """
    Recebe template, nome da aba e linha de cabecalho.
    Retorna colunas que sao de foto (Foto_X).
    """
    try:
        header_row = safe_int(header_row, 1)
        tmp_path = get_temp_filename(".xlsx")
        with open(tmp_path, "wb") as f:
            f.write(await template.read())

        wb = load_workbook(tmp_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada")

        ws = wb[sheet_name]

        columns = []
        for col_idx, cell in enumerate(ws[header_row], start=1):
            val = cell.value
            if val and is_foto_col(val):
                columns.append(
                    {
                        "name": str(val),
                        "col": col_idx,
                        "letter": get_column_letter(col_idx),
                    }
                )

        def _cleanup():
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        return JSONResponse(
            {"columns": columns},
            background=BackgroundTask(_cleanup),
        )

    except Exception as e:
        return JSONResponse(
            {"error": f"Erro ao detectar colunas Foto: {e}"},
            status_code=400,
        )


# ────────────────────────────────────────────────────────────────
#  EXPORTAR EXCEL COM IMAGENS
# ────────────────────────────────────────────────────────────────

@app.post("/api/exportar")
async def exportar_excel(
    template: UploadFile = File(...),
    sheet_name: str = Form(...),
    header_row: int = Form(...),
    assignments_json: str = Form(...),
    max_width: int = Form(300),
    max_height: int = Form(300),
    jpeg_quality: int = Form(88),
):
    """
    Recebe o template Excel + mapeamento de imagens e gera Excel com fotos.

    assignments_json esperado (exemplo):
    {
      "15": [
        {"row": 2, "dataUrl": "data:image/jpeg;base64,..."},
        {"row": 3, "dataUrl": "data:image/jpeg;base64,..."}
      ],
      "16": [
        {"row": 2, "dataUrl": "data:image/jpeg;base64,..."}
      ]
    }
    """
    try:
        header_row = safe_int(header_row, 1)
        max_width = safe_int(max_width, 300)
        max_height = safe_int(max_height, 300)
        jpeg_quality = safe_int(jpeg_quality, 88)

        try:
            assignments: Dict[str, List[Dict]] = json.loads(assignments_json or "{}")
        except Exception as e:
            raise ValueError(f"JSON de assignments inválido: {e}")

        tmp_in = get_temp_filename(".xlsx")
        with open(tmp_in, "wb") as f:
            f.write(await template.read())

        wb = load_workbook(tmp_in)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada")

        ws = wb[sheet_name]

        for col_str, items in assignments.items():
            col_idx = safe_int(col_str, 0)
            if col_idx <= 0:
                continue
            col_letter = get_column_letter(col_idx)

            for item in items:
                row = safe_int(item.get("row"), 0)
                data_url = item.get("dataUrl") or item.get("data_url")
                if row <= header_row:
                    continue
                if not data_url:
                    continue

                raw = decode_b64(data_url)
                buf, w_final, h_final = resize_image(
                    raw, max_w=max_width, max_h=max_height, quality=jpeg_quality
                )

                img = XLImage(buf)
                cell_addr = f"{col_letter}{row}"
                ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 80)
                img.anchor = cell_addr
                ws.add_image(img)

        out_path = get_temp_filename(".xlsx")
        wb.save(out_path)

        def _cleanup():
            for p in (tmp_in, out_path):
                try:
                    os.remove(p)
                except Exception:
                    pass

        filename_out = f"fotos_{os.path.basename(template.filename or 'saida')}"
        return FileResponse(
            out_path,
            filename=filename_out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(_cleanup),
        )

    except Exception as e:
        return JSONResponse({"error": f"Erro ao gerar Excel: {e}"}, status_code=400)
