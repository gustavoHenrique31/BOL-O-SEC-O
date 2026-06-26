from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import tempfile, json, os, uuid, io
from typing import List

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def norm(v):
    return str(v or '').strip().lower()


def find_photo_columns(ws, header_row=1):
    cols = []
    row  = ws[header_row]

    for cell in row:
        if cell.value and norm(cell.value).startswith('foto_'):
            try:
                int(norm(cell.value).replace('foto_', ''))
                cols.append({'name': str(cell.value).strip(), 'col': cell.column})
            except Exception:
                pass

    if cols:
        cols.sort(key=lambda x: x['col'])
        return cols

    anchor = None
    for cell in row:
        if 'tipo de quantitativo' in norm(cell.value or ''):
            anchor = cell.column
            break

    start = (anchor + 1) if anchor else 15

    for i, col in enumerate(range(start, ws.max_column + 1), 1):
        cols.append({'name': f'Foto_{i}', 'col': col})

    return cols


# ── Rota raiz — redireciona para o frontend ───────────────
@app.get("/")
def root():
    index_path = os.path.join("static", "index.html")
    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return JSONResponse({"status": "ok", "message": "API online"})


# ── API ───────────────────────────────────────────────────
@app.get("/api/")
def api_root():
    return {"status": "ok", "message": "PDF → Fotos Excel API"}


@app.post("/api/info-abas")
async def info_abas(template: UploadFile = File(...)):
    try:
        data = await template.read()
        wb   = load_workbook(io.BytesIO(data), read_only=True)
        return {"sheets": wb.sheetnames}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/info-colunas")
async def info_colunas(
    template:   UploadFile = File(...),
    sheet_name: str        = Form(...),
    header_row: int        = Form(1)
):
    try:
        data = await template.read()
        wb   = load_workbook(io.BytesIO(data))
        if sheet_name not in wb.sheetnames:
            return {"error": f"Aba '{sheet_name}' não encontrada", "columns": []}
        ws   = wb[sheet_name]
        cols = find_photo_columns(ws, header_row)
        return {"columns": cols}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/gerar-excel")
async def gerar_excel(
    template:    UploadFile       = File(...),
    images:      List[UploadFile] = File(...),
    assignments: str              = Form(...),
    sheet_name:  str              = Form(...),
    header_row:  int              = Form(1)
):
    tmp         = tempfile.mkdtemp()
    output_path = os.path.join(tmp, f"out_{uuid.uuid4().hex}.xlsx")

    try:
        

# Salvar template
        tpl_path = os.path.join(tmp, "template.xlsx")
        with open(tpl_path, "wb") as f:
            f.write(await template.read())

        

# Salvar todas as imagens com nome único
        img_map = {}
        for up in images:
            raw      = await up.read()
            img_path = os.path.join(tmp, up.filename)
            with open(img_path, "wb") as f:
                f.write(raw)
            img_map[up.filename] = img_path

        print(f"[DEBUG] Imagens recebidas: {list(img_map.keys())}")

        

# Abrir workbook
        wb = load_workbook(tpl_path)
        if sheet_name not in wb.sheetnames:
            return JSONResponse(
                {"error": f"Aba '{sheet_name}' não encontrada"},
                status_code=400
            )

        ws          = wb[sheet_name]
        photo_cols  = find_photo_columns(ws, header_row)
        col_by_name = {c['name']: c['col'] for c in photo_cols}
        assign      = json.loads(assignments)
        first_row   = header_row + 1

        print(f"[DEBUG] assignments: {assign}")
        print(f"[DEBUG] col_by_name: {col_by_name}")

        for col_name, filenames in assign.items():
            col_num = col_by_name.get(col_name)
            if col_num is None:
                print(f"[DEBUG] coluna '{col_name}' não mapeada")
                continue

            for row_offset, fname in enumerate(filenames):
                img_path = img_map.get(fname)
                if not img_path or not os.path.exists(img_path):
                    print(f"[DEBUG] imagem não encontrada: {fname}")
                    continue

                row_num = first_row + row_offset

                

# Redimensionar
                pil = PILImage.open(img_path)
                pil.thumbnail((300, 200), PILImage.LANCZOS)
                resized = os.path.join(tmp, f"r_{uuid.uuid4().hex}.jpg")
                pil.convert("RGB").save(resized, "JPEG", quality=90)

                

# Altura da linha
                _, h_px = pil.size
                ws.row_dimensions[row_num].height = max(h_px * 0.75 + 4, 20)

                

# Inserir imagem
                cell_ref      = ws.cell(row=row_num, column=col_num).coordinate
                xl_img        = XLImage(resized)
                xl_img.anchor = cell_ref
                ws.add_image(xl_img)

                print(f"[DEBUG] Inserido: {fname} → {col_name} linha {row_num}")

        wb.save(output_path)
        print(f"[DEBUG] Excel salvo: {output_path}")

        return FileResponse(
            output_path,
            filename="template_com_fotos.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[ERRO] {tb}")
        return JSONResponse({"error": str(e), "trace": tb}, status_code=500)


# ── Frontend estático — DEVE ser a última linha ───────────
app.mount("/", StaticFiles(directory="static", html=True), name="static")
